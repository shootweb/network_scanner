#!/usr/bin/env python3
"""
Internal Network Penetration Test Scanner
==========================================
Authorized use only. Run with appropriate permissions.

Usage:
  sudo python3 pentest_scanner.py                          # Auto-discover & scan entire local network
  sudo python3 pentest_scanner.py -t 192.168.1.0/24       # CIDR
  sudo python3 pentest_scanner.py -t 192.168.1.1-50       # Range
  sudo python3 pentest_scanner.py -t 192.168.1.5          # Single IP
  sudo python3 pentest_scanner.py -t 192.168.1.5,10.0.0.1/24,172.16.0.1-20  # Mixed
  sudo python3 pentest_scanner.py -t targets.txt          # File with one target per line
  sudo python3 pentest_scanner.py --no-vuln               # Skip vuln scan (faster)
  sudo python3 pentest_scanner.py -o my_report            # Custom output filename
"""

import argparse
import ipaddress
import json
import os
import re
import shutil
import subprocess
import sys
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import (Alignment, Border, Font, GradientFill,
                                  PatternFill, Side)
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[!] openpyxl not found. Attempting install...")
    # Try multiple install strategies for distro-managed Python (Kali, Ubuntu 23+, etc.)
    install_cmds = [
        [sys.executable, "-m", "pip", "install", "--break-system-packages", "openpyxl"],
        [sys.executable, "-m", "pip", "install", "openpyxl"],
        ["pip3", "install", "--break-system-packages", "openpyxl"],
        ["pip3", "install", "openpyxl"],
    ]
    installed = False
    for cmd in install_cmds:
        result = subprocess.run(cmd, capture_output=True, text=True)
        if result.returncode == 0:
            installed = True
            break
    if not installed:
        print("[!] Auto-install failed. Please run one of:")
        print("      pip install --break-system-packages openpyxl")
        print("      sudo apt install python3-openpyxl")
        sys.exit(1)
    import openpyxl
    from openpyxl.styles import (Alignment, Border, Font, GradientFill,
                                  PatternFill, Side)
    from openpyxl.utils import get_column_letter


# ─────────────────────────── Helpers ───────────────────────────

def check_nmap():
    if not shutil.which("nmap"):
        print("[!] nmap is not installed or not in PATH.")
        print("    Install with: sudo apt install nmap  /  brew install nmap")
        sys.exit(1)

def check_root():
    if os.geteuid() != 0:
        print("[!] This script requires root/sudo for SYN scans and OS detection.")
        print("    Re-run with: sudo python3 pentest_scanner.py ...")
        sys.exit(1)

def run(cmd, desc=""):
    """Run a shell command, return stdout. Streams progress to console."""
    if desc:
        print(f"    >> {desc}")
    result = subprocess.run(cmd, shell=True, capture_output=True, text=True)
    return result.stdout, result.stderr

def local_networks():
    """Detect the local subnets (/24) for all active non-loopback interfaces."""
    nets = []
    out, _ = run("ip -o -f inet addr show 2>/dev/null || ifconfig 2>/dev/null")
    # Try 'ip' output first
    for m in re.finditer(r'inet\s+(\d+\.\d+\.\d+\.\d+)/(\d+)', out):
        ip, prefix = m.group(1), int(m.group(2))
        if ip.startswith("127."):
            continue
        try:
            net = ipaddress.IPv4Network(f"{ip}/{prefix}", strict=False)
            nets.append(str(net))
        except ValueError:
            pass
    if not nets:
        # Fallback: guess 192.168.1.0/24
        nets = ["192.168.1.0/24"]
    return list(dict.fromkeys(nets))  # deduplicate


# ─────────────────────────── Target Parsing ───────────────────────────

def expand_targets(raw: str) -> list[str]:
    """
    Accept a comma-separated mix of:
      - CIDR:  192.168.1.0/24
      - Range: 192.168.1.1-50  or  10.0.0.1-10.0.0.50
      - Single IP: 192.168.1.5
      - File path: targets.txt
    Returns a flat list of individual IP strings.
    """
    targets = []
    for token in [t.strip() for t in raw.split(",") if t.strip()]:
        # File?
        if os.path.isfile(token):
            with open(token) as f:
                for line in f:
                    line = line.strip()
                    if line and not line.startswith("#"):
                        targets.extend(expand_targets(line))
            continue
        # CIDR?
        if "/" in token:
            net = ipaddress.IPv4Network(token, strict=False)
            targets.extend(str(h) for h in net.hosts())
            continue
        # Range?
        m = re.match(r'^(\d+\.\d+\.\d+\.)(\d+)-(\d+)$', token)
        if m:
            prefix, start, end = m.group(1), int(m.group(2)), int(m.group(3))
            targets.extend(f"{prefix}{i}" for i in range(start, end + 1))
            continue
        # Full IP-to-IP range: 10.0.0.1-10.0.0.50
        m2 = re.match(r'^(\d+\.\d+\.\d+\.\d+)-(\d+\.\d+\.\d+\.\d+)$', token)
        if m2:
            start_ip = ipaddress.IPv4Address(m2.group(1))
            end_ip   = ipaddress.IPv4Address(m2.group(2))
            cur = start_ip
            while cur <= end_ip:
                targets.append(str(cur))
                cur += 1
            continue
        # Single IP
        targets.append(token)
    return list(dict.fromkeys(targets))  # deduplicate, preserve order


# ─────────────────────────── Phase 1: Host Discovery ───────────────────────────

def discover_hosts(target_str: str) -> list[str]:
    """
    Ping sweep + ARP scan to find live hosts.
    target_str is passed directly to nmap (can be CIDR, space-separated IPs, etc.)
    """
    print(f"\n[*] Phase 1 — Host Discovery on: {target_str}")
    # -sn = no port scan, -PE = ICMP echo, --send-ip avoids ARP for non-local,
    # but we keep default to get ARP for local nets too.
    out, _ = run(
        f"nmap -sn --open -T4 {target_str} -oX -",
        "Running ping sweep (nmap -sn)"
    )
    live = []
    try:
        root = ET.fromstring(out)
        for host in root.findall("host"):
            status = host.find("status")
            if status is not None and status.get("state") == "up":
                addr = host.find("address[@addrtype='ipv4']")
                if addr is not None:
                    live.append(addr.get("addr"))
    except ET.ParseError:
        # Fallback: grep lines
        for m in re.finditer(r'Nmap scan report for.*?(\d+\.\d+\.\d+\.\d+)', out):
            live.append(m.group(1))
    print(f"    [+] {len(live)} host(s) found alive: {', '.join(live) or 'none'}")
    return live


# ─────────────────────────── Phase 2: Port & Service Scan ───────────────────────────

def scan_host(ip: str, run_vuln: bool, full_ports: bool = False) -> dict:
    """
    Two-pass TCP SYN scan strategy:
      Pass 1 (always): top 1000 ports, fast timing, parallel probes
      Pass 2 (--full-ports only): remaining ports after pass 1

    Service/version detection, OS fingerprinting, and optional vuln scripts
    run only on discovered open ports (--version-light keeps it quick).
    """
    scripts = "default"
    if run_vuln:
        scripts = "default,vuln"

    # Pass 1: top 1000 ports — covers ~95% of real-world services in seconds
    port_arg = "-p-" if full_ports else "--top-ports 1000"
    cmd = (
        f"nmap -sS -sV -O --version-intensity 5 "
        f"--script={scripts} "
        f"{port_arg} -T4 --open "
        f"--min-rate 1000 --max-retries 2 "
        f"{ip} -oX -"
    )
    out, err = run(cmd, f"Deep scan on {ip} ({'all ports' if full_ports else 'top 1000 ports'})")

    host_data = {
        "ip": ip,
        "hostname": "",
        "os": "",
        "os_accuracy": "",
        "ports": [],
        "vulns": [],
        "scan_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }

    try:
        root = ET.fromstring(out)
    except ET.ParseError:
        host_data["notes"] = f"XML parse error. Raw stderr: {err[:300]}"
        return host_data

    for host in root.findall("host"):
        # Hostname
        hostnames = host.find("hostnames")
        if hostnames is not None:
            hn = hostnames.find("hostname[@type='PTR']")
            if hn is None:
                hn = hostnames.find("hostname")
            if hn is not None:
                host_data["hostname"] = hn.get("name", "")

        # OS
        os_el = host.find("os")
        if os_el is not None:
            osmatch = os_el.find("osmatch")
            if osmatch is not None:
                host_data["os"] = osmatch.get("name", "")
                host_data["os_accuracy"] = osmatch.get("accuracy", "") + "%"

        # Ports
        ports_el = host.find("ports")
        if ports_el is not None:
            for port_el in ports_el.findall("port"):
                state_el = port_el.find("state")
                if state_el is None or state_el.get("state") != "open":
                    continue

                svc_el = port_el.find("service")
                port_info = {
                    "port": port_el.get("portid"),
                    "protocol": port_el.get("protocol", "tcp"),
                    "state": "open",
                    "service": svc_el.get("name", "") if svc_el is not None else "",
                    "product": svc_el.get("product", "") if svc_el is not None else "",
                    "version": svc_el.get("version", "") if svc_el is not None else "",
                    "extrainfo": svc_el.get("extrainfo", "") if svc_el is not None else "",
                    "tunnel": svc_el.get("tunnel", "") if svc_el is not None else "",
                    "scripts": [],
                }

                # Scripts on this port
                for script in port_el.findall("script"):
                    sid = script.get("id", "")
                    sout = script.get("output", "").strip()
                    port_info["scripts"].append({"id": sid, "output": sout})

                    # Collect vuln findings separately
                    if "vuln" in sid or "CVE" in sout or "VULNERABLE" in sout:
                        host_data["vulns"].append({
                            "port": port_el.get("portid"),
                            "script": sid,
                            "detail": sout[:2000],
                        })

                host_data["ports"].append(port_info)

        # Host-level scripts (e.g., smb-vuln-*)
        hostscript = host.find("hostscript")
        if hostscript is not None:
            for script in hostscript.findall("script"):
                sid = script.get("id", "")
                sout = script.get("output", "").strip()
                if "vuln" in sid or "CVE" in sout or "VULNERABLE" in sout:
                    host_data["vulns"].append({
                        "port": "host-level",
                        "script": sid,
                        "detail": sout[:2000],
                    })

    return host_data


# ─────────────────────────── Excel Report ───────────────────────────

# Color palette
C_DARK_HEADER  = "1F3864"   # Dark navy
C_MED_HEADER   = "2E75B6"   # Blue
C_LIGHT_HEADER = "D6E4F0"   # Light blue
C_GREEN        = "E2EFDA"   # Light green
C_YELLOW       = "FFF2CC"   # Light yellow
C_RED          = "FFE0E0"   # Light red
C_ORANGE       = "FCE4D6"   # Light orange
C_WHITE        = "FFFFFF"
C_VULN_RED     = "C00000"
C_ACCENT       = "00B0F0"


def style_header_cell(cell, text, bg=C_DARK_HEADER, fg=C_WHITE, size=11, bold=True, wrap=False):
    cell.value = text
    cell.font = Font(name="Arial", bold=bold, color=fg, size=size)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)


def thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)


def data_cell(cell, value, bg=None, bold=False, wrap=True, align="left"):
    cell.value = value
    cell.font = Font(name="Arial", size=10, bold=bold)
    cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
    cell.border = thin_border()
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)


def build_excel(results: list[dict], output_path: str, scan_meta: dict):
    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──
    ws_sum = wb.active
    ws_sum.title = "Executive Summary"
    _build_summary_sheet(ws_sum, results, scan_meta)

    # ── Sheet 2: Host Details ──
    ws_hosts = wb.create_sheet("Host Details")
    _build_hosts_sheet(ws_hosts, results)

    # ── Sheet 3: Open Ports ──
    ws_ports = wb.create_sheet("Open Ports")
    _build_ports_sheet(ws_ports, results)

    # ── Sheet 4: Vulnerabilities ──
    ws_vulns = wb.create_sheet("Vulnerabilities")
    _build_vulns_sheet(ws_vulns, results)

    # ── Sheet 5: Raw Script Output ──
    ws_raw = wb.create_sheet("Script Output")
    _build_raw_sheet(ws_raw, results)

    wb.save(output_path)
    print(f"\n[+] Report saved → {output_path}")


def _build_summary_sheet(ws, results, meta):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 55

    # Title block
    ws.merge_cells("A1:B1")
    c = ws["A1"]
    c.value = "🔒  Internal Penetration Test — Network Scan Report"
    c.font = Font(name="Arial", bold=True, size=16, color=C_WHITE)
    c.fill = PatternFill("solid", fgColor=C_DARK_HEADER)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:B2")
    c = ws["A2"]
    c.value = "AUTHORIZED INTERNAL USE ONLY — CONFIDENTIAL"
    c.font = Font(name="Arial", bold=True, size=10, color="C00000")
    c.fill = PatternFill("solid", fgColor="FFF2CC")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    rows = [
        ("Scan Date", meta.get("date", "")),
        ("Targets Specified", meta.get("targets", "Auto-detected local networks")),
        ("Networks Scanned", meta.get("networks", "")),
        ("Hosts Discovered", str(meta.get("hosts_found", 0))),
        ("Hosts Scanned (deep)", str(meta.get("hosts_scanned", 0))),
        ("Total Open Ports", str(meta.get("total_ports", 0))),
        ("Total Vulnerabilities", str(meta.get("total_vulns", 0))),
        ("Vuln Scan Enabled", "Yes" if meta.get("run_vuln") else "No"),
        ("Tool", "nmap + Python (pentest_scanner.py)"),
    ]

    for i, (label, val) in enumerate(rows, start=4):
        ws.row_dimensions[i].height = 20
        lc = ws.cell(row=i, column=1, value=label)
        lc.font = Font(name="Arial", bold=True, size=10, color=C_WHITE)
        lc.fill = PatternFill("solid", fgColor=C_MED_HEADER)
        lc.alignment = Alignment(horizontal="left", vertical="center")
        lc.border = thin_border()

        vc = ws.cell(row=i, column=2, value=val)
        vc.font = Font(name="Arial", size=10)
        vc.alignment = Alignment(horizontal="left", vertical="center")
        vc.border = thin_border()
        if label == "Total Vulnerabilities" and int(val or 0) > 0:
            vc.font = Font(name="Arial", size=10, bold=True, color=C_VULN_RED)

    # Host summary table
    row = len(rows) + 6
    ws.cell(row=row - 1, column=1).value = "Per-Host Summary"
    ws.cell(row=row - 1, column=1).font = Font(name="Arial", bold=True, size=11,
                                                color=C_WHITE)
    ws.cell(row=row - 1, column=1).fill = PatternFill("solid", fgColor=C_DARK_HEADER)
    ws.merge_cells(f"A{row-1}:G{row-1}")

    hdrs = ["IP Address", "Hostname", "OS", "OS Accuracy",
            "Open Ports", "Vulns Found", "Scan Time"]
    col_widths = [16, 28, 35, 12, 12, 12, 20]
    for col_letters in ["A", "B", "C", "D", "E", "F", "G"]:
        ws.column_dimensions[col_letters].width = col_widths[ord(col_letters) - 65]

    for j, h in enumerate(hdrs, 1):
        style_header_cell(ws.cell(row=row, column=j), h, bg=C_MED_HEADER)

    for rd in results:
        row += 1
        ws.row_dimensions[row].height = 18
        vuln_count = len(rd.get("vulns", []))
        cells = [
            rd.get("ip", ""),
            rd.get("hostname", ""),
            rd.get("os", "Unknown"),
            rd.get("os_accuracy", ""),
            str(len(rd.get("ports", []))),
            str(vuln_count),
            rd.get("scan_time", ""),
        ]
        for j, val in enumerate(cells, 1):
            bg = C_RED if (j == 6 and vuln_count > 0) else (
                 C_GREEN if j == 5 else None)
            data_cell(ws.cell(row=row, column=j), val, bg=bg, align="center")


def _build_hosts_sheet(ws, results):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    hdrs = ["IP Address", "Hostname", "Operating System", "OS Accuracy",
            "Open Ports #", "Services", "Vulnerabilities #", "Scan Time"]
    widths = [16, 28, 40, 12, 12, 60, 14, 20]

    ws.row_dimensions[1].height = 22
    for j, (h, w) in enumerate(zip(hdrs, widths), 1):
        ws.column_dimensions[get_column_letter(j)].width = w
        style_header_cell(ws.cell(row=1, column=j), h)

    for r, rd in enumerate(results, 2):
        ws.row_dimensions[r].height = 40
        services = "; ".join(
            f"{p['port']}/{p['protocol']} {p['service']} {p['product']} {p['version']}".strip()
            for p in rd.get("ports", [])
        )
        vuln_count = len(rd.get("vulns", []))
        row_bg = C_ORANGE if vuln_count > 0 else None

        vals = [
            rd.get("ip", ""),
            rd.get("hostname", ""),
            rd.get("os", "Unknown"),
            rd.get("os_accuracy", ""),
            len(rd.get("ports", [])),
            services,
            vuln_count,
            rd.get("scan_time", ""),
        ]
        for j, v in enumerate(vals, 1):
            data_cell(ws.cell(row=r, column=j), v, bg=row_bg)


def _build_ports_sheet(ws, results):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    hdrs = ["IP Address", "Port", "Protocol", "State", "Service",
            "Product", "Version", "Extra Info", "Tunnel/SSL"]
    widths = [16, 8, 10, 8, 16, 24, 20, 30, 12]

    ws.row_dimensions[1].height = 22
    for j, (h, w) in enumerate(zip(hdrs, widths), 1):
        ws.column_dimensions[get_column_letter(j)].width = w
        style_header_cell(ws.cell(row=1, column=j), h)

    row = 2
    for rd in results:
        for p in rd.get("ports", []):
            ws.row_dimensions[row].height = 18
            vals = [
                rd["ip"], p["port"], p["protocol"], p["state"],
                p["service"], p["product"], p["version"],
                p["extrainfo"], p["tunnel"],
            ]
            # Highlight well-known risky services
            risky = p["service"].lower() in (
                "telnet", "ftp", "rsh", "rlogin", "finger", "rexec",
                "http", "smtp", "snmp",
            )
            bg = C_YELLOW if risky else None
            for j, v in enumerate(vals, 1):
                data_cell(ws.cell(row=row, column=j), str(v) if v else "", bg=bg)
            row += 1

    if row == 2:
        ws.cell(row=2, column=1).value = "No open ports found."


def _build_vulns_sheet(ws, results):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    hdrs = ["IP Address", "Hostname", "Port", "Script / Check", "Vulnerability Detail"]
    widths = [16, 26, 14, 30, 100]

    ws.row_dimensions[1].height = 22
    for j, (h, w) in enumerate(zip(hdrs, widths), 1):
        ws.column_dimensions[get_column_letter(j)].width = w
        style_header_cell(ws.cell(row=1, column=j), h, bg=C_VULN_RED)

    row = 2
    for rd in results:
        for v in rd.get("vulns", []):
            ws.row_dimensions[row].height = 60
            vals = [rd["ip"], rd.get("hostname", ""), v["port"],
                    v["script"], v["detail"]]
            for j, val in enumerate(vals, 1):
                data_cell(ws.cell(row=row, column=j), val, bg=C_RED, wrap=True)
            row += 1

    if row == 2:
        c = ws.cell(row=2, column=1)
        c.value = "✅  No vulnerabilities detected by nmap scripts."
        c.font = Font(name="Arial", bold=True, size=11, color="00AA00")


def _build_raw_sheet(ws, results):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    hdrs = ["IP Address", "Port", "Script ID", "Output"]
    widths = [16, 8, 30, 120]

    ws.row_dimensions[1].height = 22
    for j, (h, w) in enumerate(zip(hdrs, widths), 1):
        ws.column_dimensions[get_column_letter(j)].width = w
        style_header_cell(ws.cell(row=1, column=j), h, bg=C_MED_HEADER)

    row = 2
    for rd in results:
        for p in rd.get("ports", []):
            for s in p.get("scripts", []):
                ws.row_dimensions[row].height = 45
                vals = [rd["ip"], p["port"], s["id"], s["output"]]
                for j, v in enumerate(vals, 1):
                    data_cell(ws.cell(row=row, column=j), v, wrap=True)
                row += 1

    if row == 2:
        ws.cell(row=2, column=1).value = "No script output captured."


# ─────────────────────────── Main ───────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Authorized Internal Pentest Network Scanner → Excel Report",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument(
        "-t", "--targets",
        help="Target(s): CIDR, range (x.x.x.1-50), IP, comma-mix, or file path. "
             "Omit to auto-detect local networks.",
        default=None,
    )
    parser.add_argument(
        "--no-vuln",
        action="store_true",
        help="Skip nmap vuln script category (faster scan).",
    )
    parser.add_argument(
        "--from-json",
        metavar="JSON_FILE",
        help="Skip scanning entirely and generate the Excel report from a previously saved JSON file.",
        default=None,
    )
    parser.add_argument(
        "--full-ports",
        action="store_true",
        help="Scan all 65535 ports instead of top 1000 (much slower, use on specific hosts).",
    )
    parser.add_argument(
        "-o", "--output",
        default=f"pentest_scan_{datetime.now().strftime('%Y%m%d_%H%M%S')}",
        help="Output filename prefix (no extension). Default: pentest_scan_<timestamp>",
    )
    args = parser.parse_args()

    # ── JSON-to-Excel shortcut (no scan needed, no root needed) ──
    if args.from_json:
        json_path = args.from_json
        if not os.path.isfile(json_path):
            print(f"[!] File not found: {json_path}")
            sys.exit(1)
        print(f"[*] Loading scan data from {json_path} ...")
        with open(json_path) as f:
            saved = json.load(f)
        meta    = saved.get("meta", {})
        results = saved.get("results", [])
        # Recompute totals in case meta is stale
        meta["total_ports"] = sum(len(r.get("ports", [])) for r in results)
        meta["total_vulns"] = sum(len(r.get("vulns", [])) for r in results)
        output_prefix = args.output if args.output != f"pentest_scan_{datetime.now().strftime('%Y%m%d_%H%M%S')}" else os.path.splitext(json_path)[0]
        output_xlsx = output_prefix + ".xlsx"
        print(f"[*] Loaded {len(results)} host(s)  |  {meta['total_ports']} ports  |  {meta['total_vulns']} vulns")
        print(f"[*] Generating Excel report ...")
        build_excel(results, output_xlsx, meta)
        print(f"[+] Done → {output_xlsx}")
        sys.exit(0)

    check_root()
    check_nmap()

    print("""
╔══════════════════════════════════════════════════════════╗
║      Internal Pentest Network Scanner                    ║
║      AUTHORIZED USE ONLY — CONFIDENTIAL                  ║
╚══════════════════════════════════════════════════════════╝
""")

    run_vuln = not args.no_vuln
    full_ports = args.full_ports
    date_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # ── Determine targets ──
    # IMPORTANT: never expand IPs and join into a shell string — the OS arg limit
    # (~2 MB) is exceeded on /16 or even large /24 ranges.
    # Pass original CIDR/range tokens directly to nmap (it understands them natively).
    # For Phase 2 deep scans we pass one IP at a time, so no limit is hit there.
    if args.targets:
        # Resolve any file-based target lists; keep CIDRs/ranges as-is for nmap
        raw_tokens = [t.strip() for t in args.targets.replace("\n", ",").split(",") if t.strip()]
        nmap_tokens = []
        for tok in raw_tokens:
            if os.path.isfile(tok):
                with open(tok) as f:
                    for line in f:
                        line = line.strip()
                        if line and not line.startswith("#"):
                            nmap_tokens.append(line)
            else:
                nmap_tokens.append(tok)
        nmap_target_str = " ".join(nmap_tokens)   # safe: CIDRs/ranges are short strings
        target_display  = args.targets
        networks_str    = args.targets
        # Expand only for the informational count; result is never passed to shell
        all_ips = expand_targets(args.targets)
    else:
        nets = local_networks()
        print(f"[*] No targets specified. Auto-detected local network(s): {', '.join(nets)}")
        nmap_target_str = " ".join(nets)           # CIDRs only — always short
        target_display  = "Auto-detected local network(s)"
        networks_str    = ", ".join(nets)
        all_ips = []
        for n in nets:
            net = ipaddress.IPv4Network(n, strict=False)
            all_ips.extend(str(h) for h in net.hosts())

    print(f"[*] Total IPs in scope: {len(all_ips)}")

    # ── Phase 1: Host Discovery ──
    live_hosts = discover_hosts(nmap_target_str)

    if not live_hosts:
        print("[!] No live hosts found. Exiting.")
        sys.exit(0)

    # ── Phase 2: Deep Scan ──
    print(f"\n[*] Phase 2 — Deep Scan on {len(live_hosts)} host(s)")
    print(f"    Vuln scripts : {'ENABLED' if run_vuln else 'DISABLED'}")
    print(f"    Port range   : {'ALL 65535 ports' if full_ports else 'Top 1000 (use --full-ports for all)'}")
    print("    Tip: typical scan of 6 hosts ~2-5 min. --full-ports adds 10-30 min/host.\n")

    results = []
    for i, ip in enumerate(live_hosts, 1):
        print(f"\n  [{i}/{len(live_hosts)}] Scanning {ip} ...")
        data = scan_host(ip, run_vuln, full_ports=args.full_ports)
        results.append(data)
        print(f"    Ports open: {len(data['ports'])}  |  Vulns: {len(data['vulns'])}  |  OS: {data['os'] or 'Unknown'}")

    # ── Phase 3: Build Excel ──
    total_ports = sum(len(r["ports"]) for r in results)
    total_vulns = sum(len(r["vulns"]) for r in results)

    meta = {
        "date": date_str,
        "targets": target_display,
        "networks": networks_str,
        "hosts_found": len(live_hosts),
        "hosts_scanned": len(results),
        "total_ports": total_ports,
        "total_vulns": total_vulns,
        "run_vuln": run_vuln,
    }

    output_xlsx = args.output + ".xlsx"
    output_json = args.output + ".json"

    print(f"\n[*] Phase 3 — Generating Report ...")

    # Save raw JSON as backup
    with open(output_json, "w") as f:
        json.dump({"meta": meta, "results": results}, f, indent=2)

    build_excel(results, output_xlsx, meta)

    print(f"[+] JSON backup  → {output_json}")
    print(f"\n{'='*60}")
    print(f"  Scan complete.")
    print(f"  Hosts alive:    {len(live_hosts)}")
    print(f"  Open ports:     {total_ports}")
    print(f"  Vulnerabilities:{total_vulns}")
    print(f"  Report:         {output_xlsx}")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
